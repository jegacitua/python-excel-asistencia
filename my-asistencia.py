# -*- coding: utf-8 -*-

from openpyxl import load_workbook
#from openpyxl.compat import range
from openpyxl.utils import get_column_letter

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell

import xlsxwriter

import numpy as np

import pandas as pd


#FILE_PATH = "ReporteLibroAsistencia ABRIL20.xlsx"
#SHEET = "Pagina Nº 2".decode('utf-8')

FILE_PATH = raw_input("¿Nombre del archivo:? ")
hojas = input("¿Numero de hojas del archivo excel:? ")

year = input("¿Indique el año en curso:? ")
inicio_fila = input("¿Datos inician en la fila:? ")
fin_fila = input("¿Datos finalizan en la fila:? ")




#FILE_PATH = "ReporteLibroAsistencia ABRIL20.xlsx"
#SHEET = "Pagina Nº 2".decode('utf-8')




workbook = load_workbook(FILE_PATH, read_only=True)

h=1
matriz=[]


for h in range(h,hojas+1):
    
    
    #SHEET = "Pagina N""\xc2\xba"+str(h)

    #if h == 1:
    #    SHEET = "Pagina Nº 1".decode('utf-8')
    #elif h == 2:
    #    SHEET = "Pagina Nº 2".decode('utf-8')





    #SHEET = "Pagina Nº 2".decode('utf-8')



    SHEET = "Pagina Nº "+str(h).decode('utf8')
 
    



    sheet = workbook[SHEET]

    f=inicio_fila
    
    for f in range(f,fin_fila+1):

        rut                     = ""
        funcionario             = ""
        fecha                   = ""
        entrada_programada      = ""
        salida_programada       = ""
        entrada_real            = ""
        salida_real             = ""
        horas_trabajadas        = ""
        atrasos                 = ""
        ausencias               = ""
        horas_trabajadas        = ""
        horas_permanencias      = ""
        tipo_permiso_asignado   = ""
        
        if(sheet.cell(row=f, column=1).value != "Sub-Total Semanal"):
            
            rut                     = sheet.cell(row=5, column=1).value
            funcionario             = sheet.cell(row=5, column=4).value
            fecha                   = sheet.cell(row=f, column=1).value
            entrada_programada      = sheet.cell(row=f, column=2).value      
            salida_programada       = sheet.cell(row=f, column=3).value  
            entrada_real            = sheet.cell(row=f, column=4).value  
            salida_real             = sheet.cell(row=f, column=5).value  
            horas_trabajadas        = sheet.cell(row=f, column=6).value  
            atrasos                 = sheet.cell(row=f, column=7).value  
            ausencias               = sheet.cell(row=f, column=8).value  
            horas_trabajadas        = sheet.cell(row=f, column=9).value  
            horas_permanencias      = sheet.cell(row=f, column=10).value  
            tipo_permiso_asignado   = sheet.cell(row=f, column=11).value
            
            matriz.append([str(rut.replace("R.U.T.: ","")),
                            str(funcionario.replace("Empleado: ","")),
                            str(fecha[-5:].replace("/", "-")+'-'+str(year)),
                            str(entrada_programada),
                            str(salida_programada),
                            str(entrada_real),
                            str(salida_real),
                            str(horas_trabajadas),
                            str(atrasos),
                            str(ausencias),
                            str(horas_trabajadas),
                            str(horas_permanencias),
                            str(tipo_permiso_asignado),
                            ])

                
                

print (matriz)

df = pd.DataFrame(matriz)
df.to_csv("file_path.csv")


